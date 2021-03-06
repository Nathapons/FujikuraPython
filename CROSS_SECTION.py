from tkinter import *
from tkinter import ttk
from tkinter import messagebox as msb
import openpyxl as xl
from openpyxl.styles import colors, Font
from openpyxl.drawing.image import Image
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
import xlrd
from pathlib import Path
import os
import webbrowser

class CrossSection():
    def __init__(self):
        self.link = "\\\\10.17.73.53\ORT_Result\\03.Data Cross Section\\1.Cross section\\1.For Per Day"
        self.master_link = "\\\\10.17.73.53\\ORT_Result\\03.Data Cross Section\\00.Master report"
        # self.link = 'D:\\Nathapon\\0.My work\\01.IoT\\06.VBA\\06.CROSS_SECTION\\1.For Per Day'
        # self.master_link = 'D:\\Nathapon\\0.My work\\01.IoT\\06.VBA\\06.CROSS_SECTION\\Master Report'
        self.topics_font = ('Arial', 22, 'bold')
        self.detail_font = ('Arial', 16)
        
    def main_window(self):
        root = Tk()
        WIDTH = 500
        HEIGHT = 400
        screen_width = root.winfo_screenwidth()
        screen_height = root.winfo_screenheight()
        center_width = int((screen_width/2) - (WIDTH/2))
        center_height = int((screen_height/2) - (HEIGHT/2))

        # PROGRAM PROPERTIES
        root.title("CROSS SECTION PROGRAM")
        # root.iconbitmap('fujikura_logo.ico')
        root.geometry(f'{WIDTH}x{HEIGHT}+{center_width}+{center_height}')
        root.config(bg='white')
        root.resizable(0, 0)

        # WIDGET PROPERTIES
        big_frame = Frame(root, bg='white')
        topics_label = Label(big_frame, text='CROSS SECTION PROGRAM', font=self.topics_font, fg='white', bg='blue')
        creator_label = Label(big_frame, text='Create: Nathapon.S Tel: 4308', font=self.detail_font, bg='white')
        product_frame = Frame(big_frame, bg='white')
        product_label = Label(product_frame, text='Product:', font=self.detail_font, width=8, bg='white')
        self.product_box = ttk.Combobox(product_frame, justify='center', font=self.detail_font, width=20)
        lot_frame = Frame(big_frame, bg='white')
        lot_label = Label(lot_frame, text='Lot no:', font=self.detail_font, width=8, bg='white')
        self.lot_box = ttk.Combobox(lot_frame, justify='center', font=self.detail_font, width=20)
        run_button = Button(big_frame, text='Run Program', command=self.checkUserFilled, width=15, font=('Arial', 16, 'bold'), bg='#33cc00', fg='white')
        
        # Status Report
        headers = ['TOPICS', 'STATUS']
        style = ttk.Style()
        style.configure("Treeview.Heading", font=('Arial', 10, 'bold'))
        self.status_tree = ttk.Treeview(big_frame, column=headers, show='headings', height=6)
        for header in headers:
            self.status_tree.heading(header, text=header)
            col_width = 200
            self.status_tree.column(header, anchor='center', width=col_width, minwidth=0)
        self.status_tree.tag_configure('odd', background='#E8E8E8')
        self.status_tree.tag_configure('even', background='#DFDFDF')


        # Add Event in Combobox
        self.product_box.bind('<Button-1>', self.product_box_click)
        self.lot_box.bind('<Button-1>', self.lotno_click)

        # WIDGET POSITION
        big_frame.pack()
        topics_label.grid(row=0, column=0, ipadx=5, pady=5)
        creator_label.grid(row=1, column=0, ipadx=5)
        product_frame.grid(row=2, column=0, pady=6)
        product_label.grid(row=0, column=0, padx=5)
        self.product_box.grid(row=0, column=1, padx=5, ipady=2)
        lot_frame.grid(row=3, column=0, pady=6)
        lot_label.grid(row=0, column=0, padx=5)
        self.lot_box.grid(row=0, column=1, padx=5, ipady=2)
        run_button.grid(row=4, column=0, pady=10)
        self.status_tree.grid(row=5, column=0, pady=2)

        # Activate GUI
        root.mainloop()

    def product_box_click(self, event):
        self.lot_box.set("")
        self.product_box['values'] = os.listdir(self.link)

    def lotno_click(self, event):
        product_select = self.product_box.get()

        if not product_select:
            msb.showwarning(title='Alarm to User', message='กรุณากรอกข้อมูลที่ช่อง Product')
        else:
            new_path = f'{os.path.join(self.link, product_select)}\รอ Run'

            if len(os.listdir(new_path)) > 0:
                self.lot_box['values'] = os.listdir(new_path)
            else:
                msb.showwarning(title='แจ้งเตือนไปยังผู้ใช้', message=f'ไม่มีโฟลเดอร์ที่ชื่อ รอ Run ใน {product_select}')

    def checkUserFilled(self):
        if (not self.product_box.get()) and (not self.lot_box.get()):
            msb.showwarning(title='Alarm to User', message='คุณยังกรอกข้อมูลไม่ครบถ้วน')
        else:
            self.program_overview()
    
    def program_overview(self):
        self.lotno = str(self.lot_box.get())[:9]
        result_path = f'{self.link}\{self.product_box.get()}\รอ Run\{self.lot_box.get()}\{self.lotno}.xlsx'
        report_path = f'{self.master_link}\Report_{self.product_box.get()}.xlsx'

        # Check file Exist in path?
        if os.path.isfile(result_path) and os.path.isfile(report_path):
            # Get Parameter Return Form open_excel_file function
            result_wb = self.open_excel_file_read_state(excel_path=result_path)
            report_wb = self.oepn_excel_file(excel_path=report_path)
            
            # Clear member in status_treeview
            self.clear_treeview()

            # Run function each sheetname
            report_sheets = report_wb.sheetnames
            for sheetname in report_sheets:
                report_ws = report_wb[sheetname]
                status = 'Not Import'
                
                if "OQC" in sheetname:
                    result_ws1 = result_wb.sheet_by_name('Cross Section Data')
                    result_ws2 = result_wb.sheet_by_name('Solder Mask Coverage')
                    status = self.oqc_document(report_ws, result_ws1, result_ws2)
                elif 'Solder Mask' in sheetname:
                    result_ws = result_wb.sheet_by_name('Solder Mask Coverage')
                    status = self.soldermask_document(report_ws, result_ws)
                elif 'Hot bar' in sheetname:
                    result_ws = result_wb.sheet_by_name('hotbar ')
                    status = self.hotbar_document(report_ws=report_ws, result_ws=result_ws)
                elif 'Addtion X-section and Thickness' in sheetname:
                    result_ws = result_wb.sheet_by_name('X-section')
                    status = self.addition_program(report_ws, result_ws)
                elif 'FAI' in sheetname:
                    result_ws = result_wb.sheet_by_name('FAI')
                    status = self.fai_program(report_ws, result_ws)

                data = [sheetname, status]
                self.status_tree.insert('', 'end', value=data)

            file_saveas = f'{self.link}\{self.product_box.get()}\รอ Run\{self.lot_box.get()}\Report_{self.lotno}.xlsx'
            try:
                result_wb.release_resources()
                report_wb.save(filename=file_saveas)
                # report_wb.save(filename=f'Report_{self.lotno}.xlsx')
                report_wb.close()

                # Ask user to open excel file
                ask = msb.askyesno(title="Ask to user", message=f'คุณต้องการเปิดไฟล์ {os.path.basename(file_saveas)} หรือไม่?')
                if ask:
                    webbrowser.open(file_saveas)
                else:
                    msb.showinfo(title="Ask to User", message='Import Complete')
            except Exception:
                msb.showwarning(title='Alarm to user', message=f'กรุณาปิดไฟล์ {os.path.basename(file_saveas)} แล้วรันโปรแกรมอีกครั้ง!')

        else:
            msb.showwarning(message=f'ไม่มีไฟล์ที่ชื่อ \n {result_path}', title='Alarm To User')

    def oepn_excel_file(self, excel_path):
        wb = xl.open(filename=excel_path)
        return wb

    def open_excel_file_read_state(self, excel_path):
        wb = xlrd.open_workbook(filename=excel_path)
        return wb

    def clear_treeview(self):
        for member in self.status_tree.get_children():
            self.status_tree.delete(member)

    # ------------------------------------------------ OQC Page ---------------------------------------------------------
    def oqc_document(self, report_ws, result_ws1, result_ws2):
        # result_ws1 => Cross Section Data
        # result_ws2 => Solder Mask Coverage

        # # Fill Cross Section For Stack up
        # stacks1, stacks2, stacks3, stacks4 = self.get_stack_up_data(result_ws1)
        # stack_row, stack_col = self.get_stack_up_row_col(report_ws)
        # self.fill_stackup(report_ws, stacks1, stacks2, stacks3, stacks4, stack_row, stack_col)

        # # Fill Solder mask thickness and Min PTH
        # thickness_list = self.get_solder_mask_thickness(result_ws2)
        # min_pths = self.get_min_pth_copper_thickness(result_ws1)
        # conduct_row, pth_row, fill_col = self.get_min_pth_row(report_ws)
        # self.fill_min_pth_data(report_ws, thickness_list, min_pths, conduct_row, pth_row, fill_col)

        # # Fill OQC
        # self.cross_section_for_via_and_pth(report_ws, result_ws1)
        # status = 'COMPLETE'

        try:
            # Fill Cross Section For Stack up
            stacks1, stacks2, stacks3, stacks4 = self.get_stack_up_data(result_ws1)
            stack_row, stack_col = self.get_stack_up_row_col(report_ws)
            self.fill_stackup(report_ws, stacks1, stacks2, stacks3, stacks4, stack_row, stack_col)

            # Fill Solder mask thickness and Min PTH
            thickness_list = self.get_solder_mask_thickness(result_ws2)
            min_pths = self.get_min_pth_copper_thickness(result_ws1)
            conduct_row, pth_row, fill_col = self.get_min_pth_row(report_ws)
            self.fill_min_pth_data(report_ws, thickness_list, min_pths, conduct_row, pth_row, fill_col)

            # Fill OQC
            self.cross_section_for_via_and_pth(report_ws, result_ws1)

            status = 'COMPLETE'
        except FileNotFoundError:
            status = 'Picture Error!!'
        except Exception:
            status = 'Program Error!!'

        return status

    def get_stack_up_data(self, result_ws1):
        row = 4
        stack1 = result_ws1.cell(rowx=row, colx=1).value
        stacks1 = []
        stacks2 = []
        stacks3 = []
        stacks4 = []

        while isinstance(stack1, float):
            stack2 = result_ws1.cell(rowx=row, colx=2).value
            stack3 = result_ws1.cell(rowx=row, colx=3).value
            stack4 = result_ws1.cell(rowx=row, colx=4).value
            stacks1.append(stack1)
            stacks2.append(stack2)
            stacks3.append(stack3)
            stacks4.append(stack4)

            # Loop command
            row += 1
            stack1 = result_ws1.cell(rowx=row, colx=1).value

        return stacks1, stacks2, stacks3, stacks4

    def get_stack_up_row_col(self, report_ws):
        row = 1
        max_row = report_ws.max_row
        max_col = report_ws.max_column

        while row <= max_row:
            detail1 = str(report_ws.cell(row=row, column=2).value)
            if 'CONSTRUCT' in detail1.upper():
                # ค้นหาคำว่า CPK ใน stackup
                for stack_col in range(2, max_col):
                    subdetail = str(report_ws.cell(row=row, column=stack_col).value).upper()
                    if 'CPK' in subdetail:
                        break

                # ค้นหาชื่อ stack up ตัวแรกในคอลัมน์ B
                for stack_row in range(row, row+20):
                    subdetail = str(report_ws.cell(row=stack_row, column=3).value)
                    if subdetail != 'None':
                        return stack_row, stack_col+1

            # Loop command
            row += 1

    def fill_stackup(self, report_ws, stacks1, stacks2, stacks3, stacks4, stack_row, stack_col):
        index = 0
        max_index = len(stacks1) - 1


        while index <= max_index:
            report_ws.cell(row=stack_row, column=stack_col).value = stacks1[index]
            report_ws.cell(row=stack_row, column=stack_col+1).value = stacks2[index]
            report_ws.cell(row=stack_row, column=stack_col+2).value = stacks3[index]
            report_ws.cell(row=stack_row, column=stack_col+3).value = stacks4[index]

            # Loop command
            index += 1
            stack_row += 1

    def get_solder_mask_thickness(self, result_ws2):
        row = 0
        max_row = result_ws2.nrows - 1
        thickness_list = []

        while row <= max_row:
            subdetail = str(result_ws2.cell(rowx=row, colx=1).value)
            thickness_result = str(result_ws2.cell(rowx=row, colx=2).value)

            if subdetail.upper() == 'THICKNESS' and thickness_result != "":
                for col in range(2, 12, 2):
                    thickness = result_ws2.cell(rowx=row, colx=col).value
                    thickness_list.append(thickness)

            # Loop command
            row += 1

        return thickness_list

    def get_min_pth_copper_thickness(self, result_ws1):
        min_pths = []
        row = 0
        max_row = result_ws1.nrows - 1
        min_pth1 = 7
        min_pth2 = 7
        min_pth3 = 7

        while row <= max_row:
            detail = str(result_ws1.cell(rowx=row, colx=1).value).upper()

            if 'PTH' in detail:
                for sub_row in range(row, row+15):
                    subdetail = str(result_ws1.cell(rowx=sub_row, colx=0).value).upper()
                    if 'AVE' in subdetail and (not result_ws1.cell(rowx=sub_row, colx=2).value) == False:
                        min_pth1 = round(result_ws1.cell(rowx=sub_row, colx=2).value, 3)

            elif 'OQC' in detail:
                for sub_row in range(row, max_row+1):
                    subdetail = str(result_ws1.cell(rowx=sub_row, colx=0).value).upper()
                    if 'AVE' in subdetail and (not result_ws1.cell(rowx=sub_row, colx=2).value) == False:
                        min_pth2 = round(result_ws1.cell(rowx=sub_row, colx=2).value, 3)
                        min_pth3 = round(result_ws1.cell(rowx=sub_row, colx=3).value, 3)

            # Loop command
            row += 1

        # Extend data to List
        if min_pth1 != 7 and min_pth2 != 7 and min_pth3 != 7:
            min_pths.extend((min_pth1, min_pth2, min_pth3))
            
        return min_pths

    def get_min_pth_row(self, report_ws):
        row = 1
        max_row = report_ws.max_row
        conduct_row = 1
        pth_row = 1
        fill_col = 1
        max_col = report_ws.max_column


        while row <= max_row:
            detail = str(report_ws.cell(row=row, column=2).value).upper()
            if 'CONDUCTOR' in detail:
                conduct_row = row
            elif 'PTH' in detail:
                pth_row = row

            # Loop command
            row += 1

        while fill_col <= max_col:
            detail = str(report_ws.cell(row=conduct_row - 1, column=fill_col).value)
            if 'RESULT' in detail.upper():
                break

            # Loop command
            fill_col += 1

        return conduct_row, pth_row, fill_col

    def fill_min_pth_data(self, report_ws, thickness_list, min_pths, conduct_row, pth_row, fill_col):
        index = 0
        detail = str(report_ws.cell(conduct_row-1, fill_col).value)
        
        while index < 3:
            thickness = thickness_list[index]

            report_ws.cell(conduct_row, fill_col).value = thickness
            if len(min_pths) != 0:
                min_pth = min_pths[index]
                report_ws.cell(pth_row, fill_col).value = min_pth
                report_ws.cell(pth_row, fill_col).font = Font(color='00000000')

            # Loop command
            fill_col += 1
            index += 1

    def cross_section_for_via_and_pth(self, report_ws, result_ws1):
        cross_dict = {}
        row = 0
        max_row = result_ws1.nrows - 2

        while row <= max_row:
            zone = str(result_ws1.cell(rowx=row, colx=0).value).upper()
            top = str(result_ws1.cell(rowx=row+1, colx=2).value)
            
            if zone == 'ZONE' and top != "":
                zone_no = result_ws1.cell(rowx=row, colx=1).value
                via_hole_name = str(result_ws1.cell(rowx=row-1, colx=1).value).upper()

                if via_hole_name == 'BVH':
                    via_hole_name = 'BVH'
                    last_pic_name = 'BVH'
                else:
                    via_hole_name = 'HOLE'
                    last_pic_name = 'PTH'

                zone_no_list = str(zone_no).split('-')
                bvh_zone = '==>'.join(zone_no_list)

                bottom, side_wall, adhesive = self.get_cross_section_data(result_ws1=result_ws1, cross_row=row)
                pic_path = self.get_picture_path(zone_no, last_pic_name)

                # Append value to dict
                cross_dict = {'Side': side_wall,
                              'Bottom': bottom, 
                              'Surface': top, 
                              'Etch': 'N/A', 
                              'Adhesive': adhesive, 
                              'Nickel': 'N/A',
                              'Picture': pic_path}

                fill_row = self.get_cross_section_row(report_ws, zone_no, bvh_zone, via_hole_name)
                self.fill_cross_section_data(report_ws, cross_dict, fill_row)

            # Loop command
            row += 1

    def get_cross_section_data(self, result_ws1, cross_row):
        # Assignment Value
        bottom = 0
        side_wall = 0
        adhesive = 0

        for row in range(cross_row, cross_row + 11):
            if row < result_ws1.nrows:
                detail1 = str(result_ws1.cell(rowx=row, colx=0).value).upper()
                detail2 = str(result_ws1.cell(rowx=row, colx=1).value).upper()
                result = result_ws1.cell(rowx=row, colx=2).value

                if detail2 == 'BOTTOM':
                    bottom = result
                elif 'WALL' in detail2:
                    side_wall = result
                elif 'ETCH' in detail2:
                    adhesive = result
        
        return bottom, side_wall, adhesive

    def get_picture_path(self, zone_no, last_pic_name):
        pic_path = f'{self.link}\{self.product_box.get()}\รอ Run\{self.lot_box.get()}\BVH\{zone_no} {last_pic_name}'
        file_list = os.listdir(pic_path)

        for pic_name in file_list:
            if pic_name.upper().endswith('JPG'):
                pic_name_path = os.path.join(pic_path, pic_name)
                
                return pic_name_path

    def get_cross_section_row(self, report_ws, zone_no, bvh_zone, via_hole_name):
        row = 2
        max_row = report_ws.max_row

        while row <= max_row:
            detail_a = str(report_ws.cell(row=row, column=1).value)
            cross_area_detail = str(report_ws.cell(row=row-1, column=1).value)
            bvh_detail = str(report_ws.cell(row=row-1, column=2).value)
            side_wall_value = str(report_ws.cell(row=row+1, column=2).value)

            if detail_a == 'S/N' and (cross_area_detail == zone_no or bvh_zone in bvh_detail) and side_wall_value == 'None':
                return row
            elif detail_a == 'S/N' and via_hole_name in bvh_detail.upper() and side_wall_value == 'None':
                return row

            # Loop command
            row += 1

    def fill_cross_section_data(self, report_ws, cross_dict, fill_row):
        col = 1
        max_col = report_ws.max_column

        if str(fill_row).upper() != 'NONE':
            for key in cross_dict:

                for col in range(1, 16):
                    if key.upper() in str(report_ws.cell(row=fill_row, column=col).value).upper():
                        fill_col = col
                        break

                if key != 'Picture':
                    if cross_dict[key] != 'N/A' and cross_dict[key] != '':
                        report_ws.cell(row=fill_row+1, column=fill_col).value = float(cross_dict[key])
                    else:
                        report_ws.cell(row=fill_row+1, column=fill_col).value = cross_dict[key]
                else:
                    pic = str(cross_dict['Picture'])
                    if pic.upper().endswith('.JPG'):
                        # Call add image
                        img = Image(pic)

                        p2e = pixels_to_EMU
                        c2e = cm_to_EMU
                        # Assign picture size
                        HEIGHT = 100
                        WIDTH = 100
                        # Function calculate offset
                        cellh = lambda x: c2e((x * 49.77)/99)
                        cellw = lambda x: c2e((x * (18.65-1.71))/10)
                        # Set Size and Postion
                        colloff1 = cellw(0.5)
                        rowoffset = cellh(0.5)
                        marlker = AnchorMarker(col=fill_col, colOff=colloff1, row=fill_row, rowOff=rowoffset)
                        size = XDRPositiveSize2D(p2e(HEIGHT), p2e(WIDTH))
                        # Paste Image to cell
                        img.anchor = OneCellAnchor(_from=marlker, ext=size)
                        report_ws.add_image(img)
                

    # ------------------------------------------------ Solder Mask Page -----------------------------------------------------------
    def soldermask_document(self, report_ws, result_ws):
        # flex_row, bga_pic_row, hotbar_pic_row, con_pic_row, hotbar_col, bga_col, thick_col, off_col = self.solder_get_pasteposition(report_ws=report_ws)
        # # Import data to Document
        # import_name = self.hotbar_import_data(report_ws, result_ws, flex_row, hotbar_col, bga_col, thick_col, off_col)
        # # Import picture to ducument
        # self.hotbar_import_picture(report_ws, bga_pic_row, hotbar_pic_row, con_pic_row, import_name)
        # status = 'OK'

        try:
            flex_row, bga_pic_row, hotbar_pic_row, con_pic_row, hotbar_col, bga_col, thick_col, off_col = self.solder_get_pasteposition(report_ws=report_ws)
            # Import data to Document
            import_name = self.hotbar_import_data(report_ws, result_ws, flex_row, hotbar_col, bga_col, thick_col, off_col)
            # Import picture to ducument
            status = self.hotbar_import_picture(report_ws, bga_pic_row, hotbar_pic_row, con_pic_row, import_name)
        # except FileNotFoundError:
        #     status = 'Picture Error!!'
        except Exception:
            status = 'โปรแกรมมีปัญหา!!'

        return status

    def solder_get_pasteposition(self, report_ws):
        report_row = 1
        hotbar_col = 1
        bga_col = 1
        thick_col = 1
        off_col = 1
        horbar_pic_row = 1

        while report_row != report_ws.max_row:
            detail = str(report_ws.cell(row=report_row, column=1).value).upper()

            if detail == 'FLEX NO.':
                for column in range(1, 14):
                    subdetail = str(report_ws.cell(row=report_row, column=column).value)
                    if subdetail == 'Coverage (Hot bar area)':
                        hotbar_col = column
                    elif subdetail == 'Coverage (BGA area or the nearest trace)':
                        bga_col = column
                    elif subdetail == 'solder mask thickness':
                        thick_col = column
                    elif subdetail == 'Solder mask Offset (clear from Cu)':
                        off_col = column

            elif detail == '1':
                flex_row = report_row
            elif 'BGA' in detail:
                bga_pic_row = report_row - 1
            elif 'HOT' in detail:
                hotbar_pic_row = report_row - 1
            elif 'CONNECTOR' in detail:
                con_pic_row = report_row - 1

            # Loop command
            report_row += 1

        return flex_row, bga_pic_row, hotbar_pic_row, con_pic_row, hotbar_col, bga_col, thick_col, off_col

    def hotbar_import_data(self, report_ws, result_ws, flex_row, hotbar_col, bga_col, thick_col, off_col):
        b2b_offset = result_ws.cell(rowx=6, colx=2).value
        b2b_coverage = result_ws.cell(rowx=7, colx=2).value
        b2b_thickness = result_ws.cell(rowx=8, colx=2).value

        if (not b2b_offset) == False and (not b2b_coverage) == False and (not b2b_thickness) == False:
            # import BGA to format
            import_name = 'BGA'
            result_offset_row = 6
            result_coverage_row = 7
            result_thickness_row = 8
            report_cover_column = bga_col
        else:
            # import hotbar to format
            import_name = 'HOT BAR'
            result_offset_row = 13
            result_coverage_row = 14
            result_thickness_row = 15
            report_cover_column = hotbar_col

        result_col = 2
        while result_col <= result_ws.ncols-1:
            # Get value from Result.xlsx
            left_coverage = result_ws.cell(rowx=result_coverage_row, colx=result_col).value
            right_coverage = result_ws.cell(rowx=result_coverage_row, colx=result_col+1).value
            offset = result_ws.cell(rowx=result_offset_row, colx=result_col).value
            thickness = result_ws.cell(rowx=result_thickness_row, colx=result_col).value
            # Send value to report.xlsx
            report_ws.cell(row=flex_row, column=report_cover_column).value = left_coverage
            report_ws.cell(row=flex_row, column=report_cover_column+1).value = right_coverage
            report_ws.cell(row=flex_row, column=thick_col).value = thickness
            report_ws.cell(row=flex_row, column=off_col).value = offset
            
            # Loop command
            result_col += 2
            flex_row += 1

        return import_name

    def hotbar_import_picture(self, report_ws, bga_pic_row, hotbar_pic_row, con_pic_row, import_name):
        b2b_pic_path, solder_pic_path = self.get_pic_folder()

        if import_name == 'BGA':
            row_export = bga_pic_row
        else:
            row_export = hotbar_pic_row

        # if b2b_pic_path == '' and solder_pic_path == '':
        #     status = 'ไม่มีโฟลเดอร์รูปภาพใน Solder mask'
        # elif b2b_pic_path == '':
        #     status = 'ไม่มีโฟลเดอร์ B2B ใน Solder mask'
        # elif solder_pic_path == '':
        #     status = f'ไม่มีโฟลเดอร์ {import_name} ใน Solder mask'
        # else:
        #     status = 'COMPLETE'
        
        status = 'COMPLETE'
        # Import Hot Bar& B2B Picture
        COLUMN_INSERT = 1
        if solder_pic_path != '' and len(solder_pic_path) != 0:
            solder_pics = sorted(Path(solder_pic_path).iterdir(), key=os.path.getmtime)
            for pic in solder_pics:
                if not str(pic).upper().endswith('JPG'):
                    solder_pics.remove(pic)

            for i in range(1, len(solder_pics), 2):
                if str(solder_pics[i]).upper().endswith('JPG') and str(solder_pics[i-1]).upper().endswith('JPG'):
                    # Get Picture
                    first_pic = solder_pics[i-1]
                    second_pic = solder_pics[i]
                    # Call add image
                    first_img = Image(first_pic)
                    second_img = Image(second_pic)

                    p2e = pixels_to_EMU
                    c2e = cm_to_EMU
                    # Assign picture size
                    HEIGHT = 50
                    WIDTH = 50
                    # Function calculate offset
                    cellh = lambda x: c2e((x * 49.77)/99)
                    cellw = lambda x: c2e((x * (18.65-1.71))/10)
                    # Set Size and Postion
                    colloff1 = cellw(0.1)
                    colloff2 = cellw(1)
                    rowoffset = cellh(0.5)
                    first_marker = AnchorMarker(col=COLUMN_INSERT, colOff=colloff1, row=row_export, rowOff=rowoffset)
                    second_marker = AnchorMarker(col=COLUMN_INSERT, colOff=colloff2, row=row_export, rowOff=rowoffset)
                    size = XDRPositiveSize2D(p2e(HEIGHT), p2e(WIDTH))
                    # Paste Image to cell
                    first_img.anchor = OneCellAnchor(_from=first_marker, ext=size)
                    report_ws.add_image(first_img)
                    second_img.anchor = OneCellAnchor(_from=second_marker, ext=size)
                    report_ws.add_image(second_img)
                    
                    # Prepare for next
                    COLUMN_INSERT += 1

        # Import Connector Picture
        COLUMN_INSERT = 1
        if b2b_pic_path != '' and len(b2b_pic_path) != 0:
            b2b_pics = sorted(Path(b2b_pic_path).iterdir(), key=os.path.getmtime)
            for pic in b2b_pics:
                if not str(pic).upper().endswith('JPG'):
                    b2b_pics.remove(pic)

            for i in range(1, len(b2b_pics), 2):
                if str(b2b_pics[i]).upper().endswith('JPG') and str(b2b_pics[i-1]).upper().endswith('JPG'):
                    # Get Picture
                    first_pic = b2b_pics[i-1]
                    second_pic = b2b_pics[i]
                    # Call add image
                    first_img = Image(first_pic)
                    second_img = Image(second_pic)

                    p2e = pixels_to_EMU
                    c2e = cm_to_EMU
                    # Assign picture size
                    HEIGHT = 50
                    WIDTH = 50
                    # Function calculate offset
                    cellh = lambda x: c2e((x * 49.77)/99)
                    cellw = lambda x: c2e((x * (18.65-1.71))/10)
                    # Set Size and Postion
                    colloff1 = cellw(0.1)
                    colloff2 = cellw(1)
                    rowoffset = cellh(0.5)
                    first_marker = AnchorMarker(col=COLUMN_INSERT, colOff=colloff1, row=con_pic_row, rowOff=rowoffset)
                    second_marker = AnchorMarker(col=COLUMN_INSERT, colOff=colloff2, row=con_pic_row, rowOff=rowoffset)
                    size = XDRPositiveSize2D(p2e(HEIGHT), p2e(WIDTH))
                    # Paste Image to cell
                    first_img.anchor = OneCellAnchor(_from=first_marker, ext=size)
                    report_ws.add_image(first_img)
                    second_img.anchor = OneCellAnchor(_from=second_marker, ext=size)
                    report_ws.add_image(second_img)
                    
                    # Prepare for next
                    COLUMN_INSERT += 1

        return status

    def get_pic_folder(self):
        soldermask_folder = self.get_link_soldermask_folder()
        solder_pic_path = ''
        b2b_pic_path = ''

        for pic_folder in os.listdir(soldermask_folder):
            if 'B2B' in pic_folder.upper():
                b2b_pic_path = os.path.join(soldermask_folder, pic_folder)
            elif 'HOT' in pic_folder.upper():
                solder_pic_path = os.path.join(soldermask_folder, pic_folder)
            elif 'BGA' in pic_folder.upper():
                solder_pic_path = os.path.join(soldermask_folder, pic_folder)

        return b2b_pic_path, solder_pic_path

    def get_link_soldermask_folder(self):
        folder_path = f'{self.link}\{self.product_box.get()}\รอ Run\{self.lot_box.get()}'
        folder_list = os.listdir(folder_path)

        for folder_name in folder_list:
            if 'SOLDER' in folder_name.upper():
                return os.path.join(folder_path, folder_name)

    # ---------------------------------------------------- Hot Bar Page ---------------------------------------------------------
    def hotbar_document(self, report_ws, result_ws):
        try:
            report_row, soldera_col, solderb_col, solderbstar_col = self.get_hotbar_row(report_ws=report_ws)
            report_row += 1
            result_row = 17
            total_paste = 1

            while total_paste <= 20:
                # Receive Value in cell
                solder_a = result_ws.cell(rowx=result_row, colx=2).value
                solder_b = result_ws.cell(rowx=result_row, colx=3).value
                solder_bstar = result_ws.cell(rowx=result_row, colx=4).value

                # Send value to cell
                if str(solder_a) != 'None':
                    report_ws.cell(row=report_row, column=soldera_col).value = round(solder_a, 3)

                if str(solder_b) != 'None':    
                    report_ws.cell(row=report_row, column=solderb_col).value = round(solder_b, 3)

                if str(solder_bstar) != 'None':
                    report_ws.cell(row=report_row, column=solderbstar_col).value = round(solder_bstar, 3)

                # Loop command
                result_row += 1
                report_row += 1
                total_paste += 1

            status = 'COMPLETE'
        except Exception:
            status = 'Program Error!'

        return status

    def get_hotbar_row(self, report_ws):
        report_row = 1
        column = 2

        while 'Supplier Name' not in str(report_ws.cell(row=report_row, column=column).value):
            report_row += 1

        while column != 14:
            measure_name = report_ws.cell(row=report_row, column=column).value
            if 'Solder Mask Thickness - A' in measure_name:
                soldera_col = column
                solderb_col = column + 1
                solderbstar_col = column + 2  
            # Loop command
            column += 1

        return report_row, soldera_col, solderb_col, solderbstar_col
    # ---------------------------------------------------- Addtion X-section and Thickness -----------------------------------------------
    def addition_program(self, report_ws, result_ws):
        result_col = 1
        max_col = result_ws.ncols - 1

        # while result_col < max_col:
        #     region_no = result_ws.cell(rowx=1, colx=result_col).value

        #     if isinstance(region_no, float):
        #         region_no = int(region_no)
        #         filldata_row, filldata_col, fillpic_row, fillpic_col = self.get_addition_position(report_ws, region_no)
        #         layer_dict = self.get_layer_result(result_ws, result_col)
        #         self.import_addition_data_to_report(report_ws, layer_dict, filldata_row, filldata_col)
        #         self.import_addition_pic_to_report(report_ws, region_no, fillpic_row, fillpic_col)

        #     # Prepare for next column
        #     result_col += 5

        # status = 'COMPLETE'

        try: 
            while result_col < max_col:
                region_no = result_ws.cell(rowx=1, colx=result_col).value

                if isinstance(region_no, float):
                    region_no = int(region_no)
                    filldata_row, filldata_col, fillpic_row, fillpic_col = self.get_addition_position(report_ws, region_no)
                    layer_dict = self.get_layer_result(result_ws, result_col)
                    self.import_addition_data_to_report(report_ws, layer_dict, filldata_row, filldata_col)
                    self.import_addition_pic_to_report(report_ws, region_no, fillpic_row, fillpic_col)

                # Prepare for next column
                result_col += 5

            status = 'COMPLETE'
        except FileNotFoundError:
            status = 'Picture Error!!'
        except Exception:
            status = 'Program Error!!'

        return status

    def get_addition_position(self, report_ws, region_no):
        row = 1
        filldata_row = 1
        filldata_col = 1
        fillpic_row = 1
        fillpic_col = 1
        max_row = report_ws.max_row
        max_col = report_ws.max_column + 1

        while row <= max_row:
            report_region1 = str(report_ws.cell(row=row, column=1).value).upper()
            report_region_number1 = str(report_ws.cell(row=row+1, column=1).value)
            report_region2 = str(report_ws.cell(row=row, column=2).value).upper()
            report_region_number2 = str(report_ws.cell(row=row, column=3).value).upper()
            
            if report_region1 == 'REGION' and report_region_number1 == str(region_no):
                for filldata_col in range(1, max_col):
                    subdetail = report_ws.cell(row=row, column=filldata_col).value
                    filldata_row = row + 1

                    if '#1' in subdetail:
                        return filldata_row, filldata_col, fillpic_row, fillpic_col

            elif report_region2 == 'REGION' and report_region_number2 == str(region_no):
                fillpic_row = row

                for fillpic_col in range(1, max_col):
                    subdetail = report_ws.cell(row=row, column=fillpic_col).value
                    if 'PICTURE' in str(subdetail).upper():
                        fillpic_col -= 1
                        break
                
            # Loop command
            row += 1

    def get_layer_result(self, result_ws, result_col):
        layer_dict = {}
        end_layer_col = result_col + 5
        layer_row = 4
        
        for col in range(result_col, end_layer_col):
            layer_no = result_ws.cell(rowx=2, colx=col).value

            layer_result_list = []
            while layer_row <= 12:
                layer_result = result_ws.cell(rowx=layer_row, colx=col).value
                layer_result_list.append(layer_result)

                # Loop command
                layer_row += 1
            
            # Prepare for next column
            layer_dict[layer_no] = layer_result_list
            layer_row = 4

        return layer_dict

    def import_addition_data_to_report(self, report_ws, layer_dict, filldata_row, filldata_col):
        row = filldata_row

        for layer_no in layer_dict:
            layer_results = layer_dict[layer_no]

            for layer_result in layer_results:
                report_ws.cell(row=row, column=filldata_col).value = layer_result
                row += 1
            
            # Reset row col to prepare paste next layer
            row = filldata_row
            filldata_col += 1

    def import_addition_pic_to_report(self, report_ws, region_no, fillpic_row, fillpic_col):
        addition_pic_path = self.get_addition_pic_path()
        pic_group_name = 'X' + str(region_no)
        picture_names = os.listdir(addition_pic_path)

        for picture_name in picture_names:
            picture_name_path = os.path.join(addition_pic_path, picture_name)

            if picture_name_path.upper().endswith('.JPG') and pic_group_name in picture_name:
                # Call add image
                img = Image(picture_name_path)

                p2e = pixels_to_EMU
                c2e = cm_to_EMU
                # Assign picture size
                HEIGHT = 220
                WIDTH = 250
                # Function calculate offset
                cellh = lambda x: c2e((x * 49.77)/99)
                cellw = lambda x: c2e((x * (18.65-1.71))/10)
                # Set Size and Postion
                col_offset = cellw(0.5)
                row_offset = cellh(0.75)
                first_marker = AnchorMarker(col=fillpic_col, colOff=col_offset, row=fillpic_row, rowOff=row_offset)
                size = XDRPositiveSize2D(p2e(WIDTH), p2e(HEIGHT))
                # Paste Image to cell
                img.anchor = OneCellAnchor(_from=first_marker, ext=size)
                report_ws.add_image(img)

                # Prepare for next column
                fillpic_col += 3

    def get_addition_pic_path(self):
        folder_path = f'{self.link}\{self.product_box.get()}\รอ Run\{self.lot_box.get()}'
        folder_list = os.listdir(folder_path)

        for folder_name in folder_list:
            if  folder_name.upper().startswith('X'):
                return os.path.join(folder_path, folder_name)

    # ------------------------------------------------------------------ FAI -------------------------------------------------------------------
    def fai_program(self, report_ws, result_ws):
        fai_table = self.get_data_from_result_ws(result_ws)
        report_row = self.get_fai_row(report_ws)
        self.import_fai_to_report(report_ws, fai_table, report_row)
        status = 'COMPLETE'
        # try:
        #     fai_length_list = self.get_data_from_result_ws(result_ws)
        #     report_row = self.get_fai_row(report_ws)
        #     self.import_fai_to_report(report_ws, fai_table, report_row)
        #     status = 'COMPLETE'
        # except Exception:
        #     status = 'Program Error!'

        return status

    def get_data_from_result_ws(self, result_ws):
        row = 24
        fai_table = []

        while row <= 26:
            fai_row = []
            if result_ws.cell(rowx=row, colx=1).value != "":
                for col in range(1, 7):
                    fai_length = round(result_ws.cell(rowx=row, colx=col).value, 2)
                    fai_row.append(fai_length)
                
            fai_table.append(fai_row)
            row += 1

        return fai_table

    def get_fai_row(self, report_ws):
        row = 1
        max_row = report_ws.max_row

        while row <= max_row:
            detail = report_ws.cell(row=row, column=1).value
            # Get Row Filled
            if str(detail) == '1':
                break

            # Loop command
            row += 1

        return row

    def import_fai_to_report(self, report_ws, fai_table, report_row):
        col = 2

        for fai_row in fai_table:
            row = report_row

            if len(fai_row) != 0:
                for fai_value in fai_row:
                    report_ws.cell(row=row, column=2).value = fai_value
                    row += 1

            col += 1


# Setting to Run Program
app = CrossSection()
app.main_window()